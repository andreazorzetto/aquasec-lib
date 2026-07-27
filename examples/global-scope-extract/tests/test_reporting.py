"""Tests for the report builders (Excel workbook + scope-coverage heatmap dashboard)."""
import os
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import reporting


def _result():
    return {
        "summary": {
            "repositories": {"total": 6, "scoped": 2, "unscoped": 4, "unscoped_percentage": 66.7},
            "containers": {"total": 4, "scoped": 1, "unscoped": 3, "unscoped_percentage": 75.0},
        },
        "application_scope_count": 2,
        "application_scopes": ["Team A", "Team B"],
        # master lists (Global view) that per-scope membership indexes into
        "all_repositories": [
            {"name": "nginx", "registry": "Docker Hub"},       # 0 unscoped
            {"name": "redis", "registry": "Docker Hub"},       # 1 unscoped
            {"name": "orphan/svc", "registry": "ECR"},         # 2 unscoped
            {"name": "kube-proxy", "registry": "Host Images"}, # 3 unscoped
            {"name": "scoped-a", "registry": "ECR"},           # 4 Team A & B
            {"name": "scoped-b", "registry": "ECR"},           # 5 Team B
        ],
        "all_containers": [
            {"id": "1", "name": "a", "image_name": "nginx:1", "cluster_name": "poc",
             "namespace_name": "web", "host_name": "n", "status": "running", "risk_level": ""},   # 0 unscoped
            {"id": "2", "name": "b", "image_name": "redis:7", "cluster_name": "poc",
             "namespace_name": "jobs", "host_name": "n", "status": "running", "risk_level": ""},  # 1 unscoped
            {"id": "3", "name": "c", "image_name": "envoy:2", "cluster_name": "prod",
             "namespace_name": "net", "host_name": "n", "status": "running", "risk_level": ""},   # 2 unscoped
            {"id": "9", "name": "d", "image_name": "x:1", "cluster_name": "prod",
             "namespace_name": "sys", "host_name": "n", "status": "running", "risk_level": ""},   # 3 Team A
        ],
        "scope_coverage": [
            {"scope": "(unscoped)", "unscoped": True, "repos": 4, "containers": 3,
             "repo_ids": [0, 1, 2, 3], "cont_ids": [0, 1, 2]},
            {"scope": "Team A", "repos": 1, "containers": 1, "repo_ids": [4], "cont_ids": [3]},
            {"scope": "Team B", "repos": 2, "containers": 0, "repo_ids": [4, 5], "cont_ids": []},
        ],
        "unscoped_repositories": [
            {"name": "nginx", "registry": "Docker Hub", "key": "Docker Hub/nginx"},
            {"name": "redis", "registry": "Docker Hub", "key": "Docker Hub/redis"},
            {"name": "orphan/svc", "registry": "ECR", "key": "ECR/orphan/svc"},
            {"name": "kube-proxy", "registry": "Host Images", "key": "Host Images/kube-proxy"},
        ],
        "unscoped_containers": [
            {"id": "1", "name": "a", "image_name": "nginx:1", "cluster_name": "poc",
             "namespace_name": "web", "host_name": "n", "status": "running", "risk_level": ""},
            {"id": "2", "name": "b", "image_name": "redis:7", "cluster_name": "poc",
             "namespace_name": "jobs", "host_name": "n", "status": "running", "risk_level": ""},
            {"id": "3", "name": "c", "image_name": "envoy:2", "cluster_name": "prod",
             "namespace_name": "net", "host_name": "n", "status": "running", "risk_level": ""},
        ],
    }


# --- grouping helpers ---

def test_repos_by_registry():
    assert reporting.repos_by_registry(_result()) == [
        ("Docker Hub", 2), ("ECR", 1), ("Host Images", 1)
    ]


def test_containers_by_cluster():
    assert reporting.containers_by_cluster(_result()) == [("poc", 2), ("prod", 1)]


def test_grouping_handles_missing_values():
    res = {"unscoped_repositories": [{"name": "x"}],
           "unscoped_containers": [{"id": "1", "name": "x"}]}
    assert reporting.repos_by_registry(res) == [("(none)", 1)]
    assert reporting.containers_by_cluster(res) == [("(none)", 1)]


def test_scope_coverage_sorted_pins_unscoped_first():
    rows = reporting.scope_coverage_sorted(_result(), by="repos")
    assert rows[0]["unscoped"] is True
    # non-pinned sorted by repos desc -> Team B (2) before Team A (1)
    assert [r["scope"] for r in rows[1:]] == ["Team B", "Team A"]


# --- dashboard HTML ---

def test_dashboard_is_self_contained_and_interactive():
    html = reporting.build_dashboard_html(_result(), title="Cust X", generated_at="2026-07-17 10:00")
    assert html.startswith("<!DOCTYPE html>")
    # No external references -> works offline
    assert "http://" not in html and "https://" not in html
    assert "src=" not in html
    # Title injected, data embedded, heatmap present
    assert "Cust X" in html
    assert "__DATA__" not in html
    assert "scope_coverage" in html
    assert "all_repositories" in html and "all_containers" in html
    assert "Application scope coverage" in html
    # Rows are selectable (click handler + per-row key) and drive a detail panel
    assert "data-k=" in html
    assert "#heatmap').addEventListener('click'" in html
    assert 'id="detail-title"' in html
    # Two-pane layout with a drag-to-resize splitter
    assert 'id="splitter"' in html
    assert "col-resize" in html
    # Right pane sections are explicitly labelled (chart = grouping, list = items)
    assert "Distribution by registry" in html
    assert "Distribution by cluster" in html
    # Unscoped is presented neutrally (no alarmist red treatment)
    assert "⚠ UNSCOPED" not in html
    assert "--crit" not in html
    # The old exposure filter must be gone
    assert 'data-s="exposed"' not in html
    assert "internet exposure" not in html.lower()


def test_dashboard_escapes_script_close():
    res = _result()
    res["all_repositories"].append({"name": "</script><b>x", "registry": "R"})
    html = reporting.build_dashboard_html(res)
    assert "</script><b>x" not in html
    assert "<\\/script>" in html


# --- Excel workbook ---

def test_write_xlsx_sheets_and_values(tmp_path):
    openpyxl = pytest.importorskip("openpyxl")
    path = tmp_path / "report.xlsx"
    reporting.write_xlsx(_result(), str(path), title="Cust X", generated_at="2026-07-17")
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == [
        "Summary", "Scope Coverage", "Unscoped Repositories", "By Registry",
        "Unscoped Containers", "By Cluster",
    ]
    # Scope Coverage: unscoped pinned first, then sorted by repos desc
    cov = [(r[0].value, r[1].value, r[2].value) for r in wb["Scope Coverage"].iter_rows(min_row=2)]
    assert cov[0] == ("(unscoped / Global-only)", 4, 3)
    assert cov[1][0] == "Team B" and cov[1][1] == 2
    # By Registry content
    reg = {r[0].value: r[1].value for r in wb["By Registry"].iter_rows(min_row=2)}
    assert reg == {"Docker Hub": 2, "ECR": 1, "Host Images": 1}


def test_write_xlsx_repos_only(tmp_path):
    pytest.importorskip("openpyxl")
    import openpyxl
    res = _result()
    del res["summary"]["containers"]
    del res["unscoped_containers"]
    for e in res["scope_coverage"]:
        e.pop("containers", None)
    path = tmp_path / "repos.xlsx"
    reporting.write_xlsx(res, str(path))
    wb = openpyxl.load_workbook(path)
    assert "Unscoped Containers" not in wb.sheetnames
    assert "By Cluster" not in wb.sheetnames
    assert "Unscoped Repositories" in wb.sheetnames
    assert "Scope Coverage" in wb.sheetnames
