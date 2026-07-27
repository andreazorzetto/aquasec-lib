"""
Report builders for the Global-Scope Extract utility.

Two deliverables are produced from the analysis result returned by
``aqua_global_scope_extract.analyze()``:

- ``write_xlsx()``  -- a multi-sheet Excel workbook (shareable)
- ``build_dashboard_html()`` / ``write_dashboard()`` -- a single self-contained
  HTML file (works offline, no external assets): a two-pane explorer. The left
  pane is an **application-scope coverage heatmap** (every scope as a row, bar
  length + shade encoding how many repositories / containers it covers, with the
  unscoped "no application scope" bucket pinned at the top). Clicking any row
  lists that selection's repositories (by registry) and containers (by cluster)
  in the right pane. A drag-to-resize splitter separates the two panes.

Colour encoding follows the data-viz method: magnitude -> one-hue sequential
ramps (blue for repositories, green for containers), each normalised per column.
Validated for colour-blind separation on the dark surface.

Kept separate from the CLI so the builders can be unit-tested directly.
"""

import json
from collections import Counter


# ---------------------------------------------------------------------------
# Grouping helpers (pure) -- also handy for tests
# ---------------------------------------------------------------------------

def repos_by_registry(result):
    """Return [(registry, count), ...] sorted by count desc for unscoped repos."""
    counts = Counter(r.get("registry", "") or "(none)"
                     for r in result.get("unscoped_repositories", []))
    return sorted(counts.items(), key=lambda kv: (-kv[1], kv[0]))


def containers_by_cluster(result):
    """Return [(cluster, count), ...] sorted by count desc for unscoped containers."""
    counts = Counter(c.get("cluster_name") or "(none)"
                     for c in result.get("unscoped_containers", []))
    return sorted(counts.items(), key=lambda kv: (-kv[1], kv[0]))


def scope_coverage_sorted(result, by="repos"):
    """
    Return the per-scope coverage rows sorted by a metric (desc), with the
    "(unscoped)" bucket always pinned first.
    """
    rows = result.get("scope_coverage", [])
    pinned = [r for r in rows if r.get("unscoped")]
    scopes = [r for r in rows if not r.get("unscoped")]
    scopes.sort(key=lambda r: (-r.get(by, 0), r.get("scope", "").lower()))
    return pinned + scopes


# ---------------------------------------------------------------------------
# Excel workbook
# ---------------------------------------------------------------------------

def write_xlsx(result, path, title="Unscoped Inventory (Global Scope Only)", generated_at=None):
    """
    Write the analysis result to a multi-sheet .xlsx workbook.

    Sheets: Summary, Scope Coverage, Unscoped Repositories, By Registry,
    Unscoped Containers, By Cluster. Sheets that don't apply (e.g. when
    --repos-only was used) are skipped. Requires openpyxl.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F6FEB")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    unscoped_fill = PatternFill("solid", fgColor="EAF1FB")  # neutral light-blue highlight

    wb = Workbook()

    def style_header(ws, ncols):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        if ncols:
            ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"

    def autosize(ws, headers, rows):
        for i, h in enumerate(headers, 1):
            width = len(str(h))
            for row in rows:
                width = max(width, len(str(row[i - 1])))
            ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 10), 70)

    def data_sheet(name, headers, rows, highlight=None):
        ws = wb.create_sheet(name)
        ws.append(headers)
        for row in rows:
            ws.append(list(row))
            if highlight and highlight(row):
                for col in range(1, len(headers) + 1):
                    ws.cell(row=ws.max_row, column=col).fill = unscoped_fill
        style_header(ws, len(headers))
        autosize(ws, headers, rows)
        return ws

    # --- Summary sheet (replaces default) ---
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = title
    ws["A1"].font = title_font
    r = 3
    if generated_at:
        ws.cell(row=r, column=1, value="Generated")
        ws.cell(row=r, column=2, value=generated_at)
        r += 1
    ws.cell(row=r, column=1, value="Application scopes analyzed")
    ws.cell(row=r, column=2, value=result.get("application_scope_count", 0))
    r += 2

    summary = result.get("summary", {})
    for j, h in enumerate(["Asset type", "Total", "In an app scope", "Global-only", "Global-only %"], 1):
        cell = ws.cell(row=r, column=j, value=h)
        cell.fill = header_fill
        cell.font = header_font
    r += 1
    for key, label in [("repositories", "Repositories"), ("containers", "Containers")]:
        if key in summary:
            s = summary[key]
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=s["total"])
            ws.cell(row=r, column=3, value=s["scoped"])
            ws.cell(row=r, column=4, value=s["unscoped"])
            ws.cell(row=r, column=5, value=s["unscoped_percentage"] / 100.0).number_format = "0.0%"
            r += 1
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["A"].width = 28

    # --- Scope Coverage sheet (the heatmap data) ---
    if result.get("scope_coverage"):
        cov = scope_coverage_sorted(result, by="repos")
        rows = [("(unscoped / Global-only)" if e.get("unscoped") else e["scope"],
                 e.get("repos", 0), e.get("containers", 0)) for e in cov]
        data_sheet(
            "Scope Coverage",
            ["Application scope", "Repositories", "Containers"],
            rows,
            highlight=lambda row: row[0] == "(unscoped / Global-only)",
        )

    # --- Detail sheets ---
    if "unscoped_repositories" in result:
        data_sheet(
            "Unscoped Repositories",
            ["Repository", "Registry", "Key"],
            [(x["name"], x["registry"], x["key"]) for x in result["unscoped_repositories"]],
        )
        data_sheet(
            "By Registry",
            ["Registry", "Unscoped repositories"],
            repos_by_registry(result),
        )

    if "unscoped_containers" in result:
        data_sheet(
            "Unscoped Containers",
            ["Name", "Image", "Cluster", "Namespace", "Host", "Status", "ID"],
            [(c["name"], c["image_name"], c["cluster_name"], c["namespace_name"],
              c["host_name"], c["status"], c["id"])
             for c in result["unscoped_containers"]],
        )
        data_sheet(
            "By Cluster",
            ["Cluster", "Unscoped containers"],
            containers_by_cluster(result),
        )

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Self-contained HTML dashboard
# ---------------------------------------------------------------------------

def build_dashboard_html(result, title="Unscoped Inventory — Global Scope Only", generated_at=None):
    """
    Return a single self-contained HTML string (no external assets): stat tiles +
    an interactive application-scope coverage heatmap on the left; clicking a
    scope (or the pinned Global-only bucket) lists that scope's matched
    repositories and containers on the right.
    """
    payload = {
        "title": title,
        "generated_at": generated_at or "",
        "summary": result.get("summary", {}),
        "application_scope_count": result.get("application_scope_count", 0),
        "scope_coverage": result.get("scope_coverage", []),
        "all_repositories": result.get("all_repositories", []),
        "all_containers": result.get("all_containers", []),
    }
    # </ escaped so the JSON can't close the <script> tag early.
    data_json = json.dumps(payload).replace("</", "<\\/")
    return _DASHBOARD_TEMPLATE.replace("__DATA__", data_json).replace("__TITLE__", _esc(title))


def write_dashboard(result, path, title="Unscoped Inventory — Global Scope Only", generated_at=None):
    html = build_dashboard_html(result, title=title, generated_at=generated_at)
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    return path


def _esc(s):
    return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


_DASHBOARD_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>__TITLE__</title>
<style>
/* Theme-aware chrome + validated magnitude hues (blue = repositories,
   green = containers), each a one-hue sequential ramp. Light is default;
   dark is selected via prefers-color-scheme. */
:root{
  --plane:#f9f9f7; --surface:#fcfcfb; --line:#e1e0d9; --hair:rgba(11,11,11,.10);
  --ink:#0b0b0b; --ink2:#52514e; --muted:#898781; --track:#eeede8;
  --repo:#2a78d6; --cont:#1baf7a; --accent:#2a78d6;
  --sel:rgba(42,120,214,.10); --pin:rgba(42,120,214,.06);
}
@media (prefers-color-scheme: dark){:root{
  --plane:#0d0d0d; --surface:#1a1a19; --line:#2c2c2a; --hair:rgba(255,255,255,.10);
  --ink:#fff; --ink2:#c3c2b7; --muted:#898781; --track:#26261f;
  --repo:#3987e5; --cont:#199e70; --accent:#3987e5;
  --sel:rgba(57,135,229,.16); --pin:rgba(57,135,229,.08);
}}
*{box-sizing:border-box}
body{margin:0;background:var(--plane);color:var(--ink);
  font:14px/1.5 system-ui,-apple-system,'Segoe UI',sans-serif}
header{padding:20px 24px 14px;border-bottom:1px solid var(--line)}
h1{font-size:18px;margin:0 0 3px}
.sub{color:var(--ink2);font-size:12.5px}
.tiles{display:flex;flex-wrap:wrap;gap:12px;padding:16px 24px}
.tile{background:var(--surface);border:1px solid var(--hair);border-radius:10px;
  padding:12px 16px;min-width:150px;flex:1}
.tile .n{font-size:24px;font-weight:700;font-variant-numeric:tabular-nums}
.tile .l{color:var(--ink2);font-size:11px;text-transform:uppercase;letter-spacing:.05em}

/* two-pane layout with a drag-to-resize splitter */
.wrap{display:flex;align-items:stretch;padding:8px 24px 28px}
.card{background:var(--surface);border:1px solid var(--hair);border-radius:12px;
  padding:16px 18px;min-width:0}                 /* min-width:0 => long names can't blow out the pane */
#paneL{flex:0 0 46%}
#detail{flex:1 1 auto}
.splitter{flex:0 0 16px;align-self:stretch;cursor:col-resize;position:relative}
.splitter::before{content:"";position:absolute;left:7px;top:50%;transform:translateY(-50%);
  width:2px;height:44px;border-radius:2px;background:var(--line)}
.splitter:hover::before,.splitter.drag::before{background:var(--accent);height:64px}
@media(max-width:940px){
  .wrap{flex-direction:column}
  #paneL{flex:1 1 auto!important}
  .splitter{display:none}
  #detail{margin-top:16px}
}

.card h2{font-size:14px;margin:0 0 4px}
.card .note{color:var(--muted);font-size:11px;margin:0 0 12px}
.controls{display:flex;gap:8px;flex-wrap:wrap;margin:0 0 10px;align-items:center}
input.search,select{background:var(--plane);border:1px solid var(--line);color:var(--ink);
  border-radius:7px;padding:6px 9px;font-size:12.5px;outline:none;min-width:0}
input.search{flex:1;min-width:120px}
input.search:focus,select:focus{border-color:var(--accent)}

/* heatmap rows */
.hm{max-height:560px;overflow:auto}
.hrow{display:grid;grid-template-columns:1fr 92px 92px;gap:10px;align-items:center;
  padding:4px 6px;border-bottom:1px solid var(--line);font-size:12.5px;cursor:pointer;border-radius:6px}
.hrow:hover{background:var(--sel)}
.hrow.sel{outline:2px solid var(--accent);outline-offset:-2px;background:var(--sel)}
.hrow.head{position:sticky;top:0;background:var(--surface);color:var(--muted);cursor:default;
  font-size:10.5px;text-transform:uppercase;letter-spacing:.05em;border-bottom:1px solid var(--line);z-index:1}
.hrow.head:hover{background:var(--surface)}
.hrow.head .cell{justify-content:flex-end;display:flex}
.hrow.pin{background:var(--pin);border-bottom:1px solid var(--line)}   /* the unscoped bucket: pinned, neutral */
.hrow .name{white-space:nowrap;overflow:hidden;text-overflow:ellipsis;min-width:0}
.hrow .name .dot{display:inline-block;width:6px;height:6px;border-radius:50%;background:var(--muted);
  margin-right:6px;vertical-align:1px}
.cell{display:grid;grid-template-columns:1fr auto;gap:6px;align-items:center;min-width:0}
.hbar{height:14px;background:var(--track);border-radius:4px;overflow:hidden}
.hbar>i{display:block;height:100%;border-radius:4px}
.cell .v{font-variant-numeric:tabular-nums;color:var(--ink2);min-width:28px;text-align:right}
.legend{display:flex;gap:14px;color:var(--muted);font-size:11px;margin-top:8px;flex-wrap:wrap;align-items:center}
.legend .sw{width:10px;height:10px;border-radius:2px;display:inline-block;margin-right:4px;vertical-align:-1px}

/* detail (right): two clearly-delimited sections, each with a chart + a list */
#detail h2 .tag{font-weight:400;color:var(--muted);font-size:12px}
.sec{border:1px solid var(--line);border-radius:10px;background:var(--plane);padding:12px 14px;margin-top:12px}
.sec-head{display:flex;justify-content:space-between;align-items:baseline;margin-bottom:2px}
.sec-title{font-size:13px;font-weight:600}
.sec-title .sw{width:9px;height:9px;border-radius:2px;display:inline-block;margin-right:7px;vertical-align:1px}
.sec-count{color:var(--muted);font-size:12px;font-variant-numeric:tabular-nums}
.block{margin-top:12px}
.block-lbl{font-size:10.5px;text-transform:uppercase;letter-spacing:.05em;color:var(--muted);
  margin:0 0 7px;display:flex;justify-content:space-between;align-items:center;gap:10px}
.block-lbl .search{flex:0 1 190px}
.bars{display:flex;flex-direction:column;gap:6px}
.barrow{display:grid;grid-template-columns:150px 1fr 40px;gap:10px;align-items:center;font-size:12px}
.barrow .lab{white-space:nowrap;overflow:hidden;text-overflow:ellipsis;min-width:0}
.barrow .track{height:14px;background:var(--track);border-radius:4px;overflow:hidden}
.barrow .track>i{display:block;height:100%;border-radius:4px}
.barrow .val{text-align:right;color:var(--ink2);font-variant-numeric:tabular-nums}
.list{max-height:240px;overflow:auto;margin-top:2px}
.row{display:flex;gap:8px;align-items:center;padding:4px 2px;border-bottom:1px solid var(--line);font-size:12px}
.row .nm{flex:1;min-width:0;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.row .meta{color:var(--muted);font-size:11px;white-space:nowrap;max-width:45%;overflow:hidden;text-overflow:ellipsis}
.empty{color:var(--muted);font-size:12.5px;padding:8px 2px}
footer{color:var(--muted);font-size:11px;padding:0 24px 24px}
#tip{position:fixed;pointer-events:none;background:var(--surface);border:1px solid var(--line);
  border-radius:6px;padding:6px 9px;font-size:11.5px;color:var(--ink);box-shadow:0 4px 14px rgba(0,0,0,.25);
  opacity:0;transition:opacity .08s;z-index:20;max-width:280px}
</style></head><body>
<div id="tip"></div>
<header>
  <h1 id="title"></h1>
  <div class="sub" id="subtitle"></div>
</header>
<div class="tiles" id="tiles"></div>
<div class="wrap" id="wrap">
  <div class="card" id="paneL">
    <h2>Application scope coverage</h2>
    <p class="note">Repositories and containers each scope covers (bar length &amp; shade ∝ count, per column). The unscoped bucket (assets in no application scope) is pinned at the top. <b>Click a row</b> to list its resources →</p>
    <div class="controls">
      <input class="search" id="scopesearch" placeholder="filter scopes…" autocomplete="off">
      <select id="sortsel" title="sort scopes by">
        <option value="repos">sort: repositories</option>
        <option value="containers">sort: containers</option>
        <option value="scope">sort: name</option>
      </select>
    </div>
    <div class="hm" id="heatmap"></div>
    <div class="legend">
      <span><span class="sw" style="background:var(--repo)"></span>repositories</span>
      <span><span class="sw" style="background:var(--cont)"></span>containers</span>
      <span>· top row = unscoped (no application scope)</span>
    </div>
  </div>
  <div class="splitter" id="splitter" title="drag to resize"></div>
  <div class="card" id="detail">
    <h2 id="detail-title"></h2>
    <p class="note" id="detail-sub"></p>

    <div class="sec" id="sec-repo">
      <div class="sec-head">
        <span class="sec-title"><span class="sw" style="background:var(--repo)"></span>Repositories</span>
        <span class="sec-count" id="repo-c"></span>
      </div>
      <div class="block">
        <div class="block-lbl">Distribution by registry</div>
        <div class="bars" id="repobars"></div>
      </div>
      <div class="block">
        <div class="block-lbl"><span>Repository list</span>
          <input class="search" id="reposearch" placeholder="filter…" autocomplete="off"></div>
        <div class="list" id="repolist"></div>
      </div>
    </div>

    <div class="sec" id="sec-cont">
      <div class="sec-head">
        <span class="sec-title"><span class="sw" style="background:var(--cont)"></span>Containers</span>
        <span class="sec-count" id="cont-c"></span>
      </div>
      <div class="block">
        <div class="block-lbl">Distribution by cluster</div>
        <div class="bars" id="contbars"></div>
      </div>
      <div class="block">
        <div class="block-lbl"><span>Container list</span>
          <input class="search" id="contsearch" placeholder="filter…" autocomplete="off"></div>
        <div class="list" id="contlist"></div>
      </div>
    </div>
  </div>
</div>
<footer id="foot"></footer>
<script>
const D = __DATA__;
const $ = s => document.querySelector(s);
const fmt = n => (n||0).toLocaleString();
const esc = s => String(s==null?'':s).replace(/[&<>]/g,m=>({'&':'&amp;','<':'&lt;','>':'&gt;'}[m]));
const REPOS = D.all_repositories||[], CONTS = D.all_containers||[];
const COV = D.scope_coverage||[];
const sr = D.summary.repositories, sc = D.summary.containers;
const hasRepos = !!sr, hasCont = !!sc;
const CAP = 400;
const tip = $('#tip');
function showTip(html,e){tip.innerHTML=html;tip.style.opacity=1;moveTip(e);}
function moveTip(e){const p=12;let x=e.clientX+p,y=e.clientY+p;
  if(x+tip.offsetWidth>innerWidth)x=e.clientX-tip.offsetWidth-p;
  if(y+tip.offsetHeight>innerHeight)y=e.clientY-tip.offsetHeight-p;
  tip.style.left=x+'px';tip.style.top=y+'px';}
function hideTip(){tip.style.opacity=0;}

$('#title').textContent = D.title; document.title = D.title;
$('#subtitle').textContent = (D.generated_at?('Generated '+D.generated_at+' · '):'')
  + D.application_scope_count + ' application scope(s) analyzed';

// ---- stat tiles (neutral; the numbers speak for themselves) ----
const tiles=[];
if(sr){tiles.push({n:fmt(sr.total),l:'Repositories total'});
  tiles.push({n:fmt(sr.unscoped)+' ('+sr.unscoped_percentage.toFixed(1)+'%)',l:'Repos — no app scope'});}
if(sc){tiles.push({n:fmt(sc.total),l:'Containers total'});
  tiles.push({n:fmt(sc.unscoped)+' ('+sc.unscoped_percentage.toFixed(1)+'%)',l:'Containers — no app scope'});}
$('#tiles').innerHTML=tiles.map(t=>`<div class="tile"><div class="n">${t.n}</div><div class="l">${esc(t.l)}</div></div>`).join('');

// ---- heatmap ----
const maxRepos=Math.max(1,...COV.map(r=>r.repos||0));
const maxCont =Math.max(1,...COV.map(r=>r.containers||0));
const totRepos=sr?sr.total:0, totCont=sc?sc.total:0;
let sortKey='repos', scopeQuery='', selKey=null;
function keyOf(e){return e.unscoped?'(unscoped)':e.scope;}
function heatCell(v,max,hue){
  if(!v) return `<div class="cell"><div class="hbar"></div><span class="v">0</span></div>`;
  const w=Math.max(3,100*v/max), a=(0.35+0.65*v/max).toFixed(2);
  return `<div class="cell"><div class="hbar"><i style="width:${w.toFixed(1)}%;background:${hue};opacity:${a}"></i></div><span class="v">${fmt(v)}</span></div>`;
}
function renderHeatmap(){
  const pinned=COV.filter(r=>r.unscoped);
  let scopes=COV.filter(r=>!r.unscoped);
  if(scopeQuery) scopes=scopes.filter(r=>r.scope.toLowerCase().includes(scopeQuery));
  if(sortKey==='scope') scopes.sort((a,b)=>a.scope.toLowerCase().localeCompare(b.scope.toLowerCase()));
  else scopes.sort((a,b)=>(b[sortKey]||0)-(a[sortKey]||0)||a.scope.toLowerCase().localeCompare(b.scope.toLowerCase()));
  const rows=pinned.concat(scopes);
  const head=`<div class="hrow head"><span class="name">application scope (${scopes.length})</span>`
    +(hasRepos?`<span class="cell">repos</span>`:``)+(hasCont?`<span class="cell">containers</span>`:``)+`</div>`;
  $('#heatmap').innerHTML=head+rows.map(r=>{
    const k=keyOf(r);
    const label=r.unscoped?'<span class="dot"></span>Unscoped (no application scope)':esc(r.scope);
    return `<div class="hrow ${r.unscoped?'pin':''} ${k===selKey?'sel':''}" data-k="${esc(k)}" data-r="${r.repos}" data-c="${r.containers}" data-u="${r.unscoped?1:0}">`
      +`<span class="name" title="${esc(r.unscoped?'Unscoped (no application scope)':r.scope)}">${label}</span>`
      +(hasRepos?heatCell(r.repos,maxRepos,'var(--repo)'):``)
      +(hasCont?heatCell(r.containers,maxCont,'var(--cont)'):``)+`</div>`;
  }).join('');
}
$('#heatmap').addEventListener('mousemove',e=>{
  const row=e.target.closest('.hrow'); if(!row||row.classList.contains('head')){hideTip();return;}
  const nm=row.dataset.u==='1'?'Unscoped (no application scope)':row.dataset.k, rr=+row.dataset.r, cc=+row.dataset.c;
  let h=`<b>${esc(nm)}</b>`;
  if(hasRepos)h+=`<br>${fmt(rr)} repositories`+(totRepos?` · ${(100*rr/totRepos).toFixed(1)}% of all`:``);
  if(hasCont)h+=`<br>${fmt(cc)} containers`+(totCont?` · ${(100*cc/totCont).toFixed(1)}% of all`:``);
  h+=`<br><span style="color:var(--muted)">click to list resources</span>`;
  showTip(h,e);
});
$('#heatmap').addEventListener('mouseleave',hideTip);
$('#heatmap').addEventListener('click',e=>{
  const row=e.target.closest('.hrow'); if(!row||row.classList.contains('head'))return;
  select(row.dataset.k);
});
$('#scopesearch').addEventListener('input',e=>{scopeQuery=e.target.value.toLowerCase();renderHeatmap();});
$('#sortsel').addEventListener('change',e=>{sortKey=e.target.value;renderHeatmap();});

// ---- detail (right) driven by selection ----
function groupCount(items,keyFn){const m=new Map();for(const it of items){const k=(keyFn(it)||'(none)');m.set(k,(m.get(k)||0)+1);}
  return [...m.entries()].sort((a,b)=>b[1]-a[1]||String(a[0]).localeCompare(String(b[0])));}
function miniBars(el,pairs,hue){
  const max=Math.max(1,...pairs.map(p=>p[1]));
  el.innerHTML = pairs.length ? pairs.slice(0,12).map(([lab,n])=>{
    const w=Math.max(3,100*n/max), a=(0.35+0.65*n/max).toFixed(2);
    return `<div class="barrow"><span class="lab" title="${esc(lab)}">${esc(lab)}</span>`
      +`<span class="track"><i style="width:${w.toFixed(1)}%;background:${hue};opacity:${a}"></i></span>`
      +`<span class="val">${fmt(n)}</span></div>`;
  }).join('') : '<div class="empty">none</div>';
}
let repoQ='', contQ='', selRepos=[], selConts=[];
function renderRepoList(){
  const q=repoQ.toLowerCase();
  const rows=q?selRepos.filter(r=>((r.name||'')+' '+(r.registry||'')).toLowerCase().includes(q)):selRepos;
  const shown=rows.slice(0,CAP);
  $('#repolist').innerHTML = rows.length ? shown.map(r=>`<div class="row"><span class="nm" title="${esc(r.name)}">${esc(r.name)}</span><span class="meta" title="${esc(r.registry)}">${esc(r.registry)}</span></div>`).join('')
    + (rows.length>CAP?`<div class="empty">+${fmt(rows.length-CAP)} more — type to filter</div>`:``)
    : '<div class="empty">No matching repositories</div>';
}
function renderContList(){
  const q=contQ.toLowerCase();
  const rows=q?selConts.filter(c=>((c.name||'')+' '+(c.image_name||'')+' '+(c.cluster_name||'')+' '+(c.namespace_name||'')).toLowerCase().includes(q)):selConts;
  const shown=rows.slice(0,CAP);
  $('#contlist').innerHTML = rows.length ? shown.map(c=>{const m=`${esc(c.image_name||'—')} · ${esc(c.cluster_name||'—')}/${esc(c.namespace_name||'—')}`;
    return `<div class="row"><span class="nm" title="${esc(c.name)}">${esc(c.name)}</span><span class="meta" title="${m}">${m}</span></div>`;}).join('')
    + (rows.length>CAP?`<div class="empty">+${fmt(rows.length-CAP)} more — type to filter</div>`:``)
    : '<div class="empty">No matching containers</div>';
}
function select(k){
  const entry=COV.find(e=>keyOf(e)===k) || COV[0];
  selKey=keyOf(entry);
  try{location.hash = entry.unscoped?'#unscoped':('#scope='+encodeURIComponent(entry.scope));}catch(e){}
  selRepos=(entry.repo_ids||[]).map(i=>REPOS[i]).filter(Boolean);
  selConts=(entry.cont_ids||[]).map(i=>CONTS[i]).filter(Boolean);
  const nm = entry.unscoped?'Unscoped':esc(entry.scope);
  const tag = entry.unscoped?' — not in any application scope':' — application scope';
  $('#detail-title').innerHTML = nm+`<span class="tag">${tag}</span>`;
  $('#detail-sub').textContent = `${fmt(selRepos.length)} repositories · ${fmt(selConts.length)} containers`;
  $('#repo-c').textContent = fmt(selRepos.length);
  $('#cont-c').textContent = fmt(selConts.length);
  if(hasRepos) miniBars($('#repobars'), groupCount(selRepos,r=>r.registry), 'var(--repo)');
  if(hasCont)  miniBars($('#contbars'), groupCount(selConts,c=>c.cluster_name), 'var(--cont)');
  renderRepoList(); renderContList();
  document.querySelectorAll('.hrow').forEach(row=>row.classList.toggle('sel', row.dataset.k===selKey));
}
$('#reposearch').addEventListener('input',e=>{repoQ=e.target.value;renderRepoList();});
$('#contsearch').addEventListener('input',e=>{contQ=e.target.value;renderContList();});

// ---- draggable splitter between the two panes ----
(function(){
  const wrap=$('#wrap'), splitter=$('#splitter'), paneL=$('#paneL');
  let dragging=false;
  splitter.addEventListener('mousedown',e=>{dragging=true;splitter.classList.add('drag');
    document.body.style.userSelect='none';document.body.style.cursor='col-resize';e.preventDefault();});
  window.addEventListener('mousemove',e=>{
    if(!dragging)return;
    const rect=wrap.getBoundingClientRect();
    let w=e.clientX-rect.left;
    const min=280, max=rect.width-340;
    w=Math.max(min,Math.min(Math.max(min,max),w));
    paneL.style.flex='0 0 '+w+'px';
  });
  window.addEventListener('mouseup',()=>{if(dragging){dragging=false;splitter.classList.remove('drag');
    document.body.style.userSelect='';document.body.style.cursor='';}});
})();

// ---- init + deep link (#unscoped or #scope=Name) ----
function initSelectionFromHash(){
  const h=decodeURIComponent(location.hash.replace(/^#/,''));
  if(h.startsWith('scope=')){const nm=h.slice(6); if(COV.some(e=>e.scope===nm)) return nm;}
  return '(unscoped)';
}
renderHeatmap();
select(initSelectionFromHash());
$('#foot').textContent='Aqua application-scope coverage · click any scope to list its repositories and containers; the unscoped bucket is what no application scope selects.';
</script>
</body></html>
"""
